import asyncio
import importlib
import inspect
import json
import os
import subprocess
import sys
import traceback
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any, Callable, Optional


class SkillCategory(str, Enum):
    REASONING = "reasoning"
    VISION = "vision"
    CODE = "code"
    OFFICE = "office"
    BROWSER = "browser"
    GUI = "gui"
    DATA = "data"
    COMMUNICATION = "communication"
    CUSTOM = "custom"


class SkillLevel(str, Enum):
    BASIC = "basic"
    INTERMEDIATE = "intermediate"
    ADVANCED = "advanced"
    EXPERT = "expert"


@dataclass
class SkillManifest:
    name: str
    version: str = "0.1.0"
    category: SkillCategory = SkillCategory.CUSTOM
    level: SkillLevel = SkillLevel.BASIC
    description: str = ""
    author: str = ""
    dependencies: list[str] = field(default_factory=list)
    entry_point: str = ""
    requires_auth: bool = False
    requires_gpu: bool = False
    requires_browser: bool = False
    is_streaming: bool = False
    tags: list[str] = field(default_factory=list)
    metadata: dict = field(default_factory=dict)


class BaseSkill(ABC):
    manifest: SkillManifest

    @abstractmethod
    async def execute(self, **kwargs) -> dict:
        ...

    async def validate(self) -> bool:
        return True

    async def setup(self):
        pass

    async def teardown(self):
        pass

    def describe(self) -> dict:
        return {
            "name": self.manifest.name,
            "category": self.manifest.category.value,
            "level": self.manifest.level.value,
            "description": self.manifest.description,
            "tags": self.manifest.tags,
        }


class ReasoningSkill(BaseSkill):
    category = SkillCategory.REASONING

    def __init__(self):
        self.manifest = SkillManifest(
            name="reasoning-chain",
            version="0.1.0",
            category=SkillCategory.REASONING,
            level=SkillLevel.ADVANCED,
            description="PanguAgent-style intrinsic/extrinsic reasoning chain with tree search",
            tags=["reasoning", "pangu", "tree-search", "planning"],
        )
        self._reasoning_chain = None

    async def setup(self):
        try:
            from cogu.core.reasoning_chain import ReasoningChain, ChainContext
            self._reasoning_chain = ReasoningChain
        except ImportError:
            pass

    async def execute(self, **kwargs) -> dict:
        mode = kwargs.get("mode", "react")
        query = kwargs.get("query", "")
        context = kwargs.get("context", {})
        max_iterations = kwargs.get("max_iterations", 10)

        result = {
            "mode": mode,
            "query": query,
            "iterations": 0,
            "conclusion": "",
            "trace": [],
        }

        try:
            from cogu.core.reasoning_chain import ChainContext

            ctx = ChainContext(initial_state={"query": query, **context})

            if mode == "swift_sage":
                from cogu.core.reasoning_chain import SwiftSageReasoning
                reasoner = SwiftSageReasoning()
                ctx = await reasoner.execute(ctx)
            elif mode == "reflect":
                from cogu.core.reasoning_chain import ReflectReasoning
                reasoner = ReflectReasoning()
                ctx = await reasoner.execute(ctx)
            else:
                from cogu.core.reasoning_chain import ReactReasoning
                reasoner = ReactReasoning()
                ctx = await reasoner.execute(ctx)

            result["conclusion"] = ctx.final_answer or str(ctx.state)
            result["iterations"] = ctx.iteration_count
            result["trace"] = ctx.iteration_log

        except ImportError:
            result["conclusion"] = f"[ReasoningChain not available] Query: {query}"
        except Exception as e:
            result["error"] = str(e)
            result["conclusion"] = f"Reasoning error: {e}"

        return result


class GuiVisionSkill(BaseSkill):
    category = SkillCategory.VISION

    def __init__(self):
        self.manifest = SkillManifest(
            name="gui-vision",
            version="0.1.0",
            category=SkillCategory.VISION,
            level=SkillLevel.ADVANCED,
            description="GUI visual operation: screenshot analysis, element detection, click automation",
            tags=["vision", "gui", "screenshot", "automation", "computer-use"],
            requires_gpu=False,
        )
        self._screenshot_dir: Optional[Path] = None

    async def setup(self):
        self._screenshot_dir = Path(os.environ.get("COGU_SCREENSHOT_DIR", Path.home() / ".cogu" / "screenshots"))
        self._screenshot_dir.mkdir(parents=True, exist_ok=True)

    async def execute(self, **kwargs) -> dict:
        action = kwargs.get("action", "analyze")
        result = {"action": action, "success": False}

        if action == "screenshot":
            result.update(await self._take_screenshot(**kwargs))
        elif action == "analyze_region":
            result.update(await self._analyze_region(**kwargs))
        elif action == "click":
            result.update(await self._click_element(**kwargs))
        elif action == "type_text":
            result.update(await self._type_text(**kwargs))
        elif action == "detect_elements":
            result.update(await self._detect_elements(**kwargs))
        elif action == "read_screen":
            result.update(await self._read_screen(**kwargs))
        else:
            result["error"] = f"Unknown action: {action}"

        return result

    async def _take_screenshot(self, **kwargs) -> dict:
        region = kwargs.get("region")
        output_path = kwargs.get("output_path")
        if not output_path:
            import time
            output_path = str(self._screenshot_dir / f"screenshot_{int(time.time())}.png")

        try:
            if sys.platform == "win32":
                import ctypes
                from ctypes import wintypes
                user32 = ctypes.windll.user32
                gdi32 = ctypes.windll.gdi32
                width = user32.GetSystemMetrics(0)
                height = user32.GetSystemMetrics(1)
                if region:
                    x, y, w, h = region
                else:
                    x, y, w, h = 0, 0, width, height
                hdc_screen = user32.GetDC(0)
                hdc_mem = gdi32.CreateCompatibleDC(hdc_screen)
                hbmp = gdi32.CreateCompatibleBitmap(hdc_screen, w, h)
                gdi32.SelectObject(hdc_mem, hbmp)
                gdi32.BitBlt(hdc_mem, 0, 0, w, h, hdc_screen, x, y, 0x00CC0020)
                self._save_bitmap(hbmp, output_path, w, h)
                gdi32.DeleteObject(hbmp)
                gdi32.DeleteDC(hdc_mem)
                user32.ReleaseDC(0, hdc_screen)
            elif sys.platform == "darwin":
                cmd = ["screencapture", "-x"]
                if region:
                    x, y, w, h = region
                    cmd.extend(["-R", f"{x},{y},{w},{h}"])
                cmd.append(output_path)
                subprocess.run(cmd, check=True, capture_output=True)
            else:
                subprocess.run(["import", "-window", "root", output_path], check=True, capture_output=True)

            return {"success": True, "output_path": output_path, "width": w if region else width, "height": h if region else height}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def _save_bitmap(self, hbmp, path, width, height):
        import ctypes
        import struct
        gdi32 = ctypes.windll.gdi32
        class BITMAPINFOHEADER(ctypes.Structure):
            _fields_ = [
                ("biSize", ctypes.c_uint32), ("biWidth", ctypes.c_int32), ("biHeight", ctypes.c_int32),
                ("biPlanes", ctypes.c_uint16), ("biBitCount", ctypes.c_uint16), ("biCompression", ctypes.c_uint32),
                ("biSizeImage", ctypes.c_uint32), ("biXPelsPerMeter", ctypes.c_int32), ("biYPelsPerMeter", ctypes.c_int32),
                ("biClrUsed", ctypes.c_uint32), ("biClrImportant", ctypes.c_uint32),
            ]
        bi = BITMAPINFOHEADER()
        bi.biSize = ctypes.sizeof(BITMAPINFOHEADER)
        bi.biWidth = width
        bi.biHeight = -height
        bi.biPlanes = 1
        bi.biBitCount = 32
        bi.biCompression = 0
        buf = ctypes.create_string_buffer(width * height * 4)
        gdi32.GetDIBits(gdi32.CreateCompatibleDC(0), hbmp, 0, height, buf, ctypes.byref(bi), 0)
        with open(path, "wb") as f:
            f.write(b"BM")
            file_size = 54 + len(buf)
            f.write(struct.pack("<I", file_size))
            f.write(b"\x00\x00\x00\x00\x36\x00\x00\x00\x28\x00\x00\x00")
            f.write(struct.pack("<i", width))
            f.write(struct.pack("<i", height))
            f.write(b"\x01\x00\x20\x00\x00\x00\x00\x00")
            f.write(struct.pack("<I", len(buf)))
            f.write(b"\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00")
            f.write(buf.raw)

    async def _analyze_region(self, **kwargs) -> dict:
        return {"success": True, "elements": [], "message": "Region analysis requires LLM vision model integration"}

    async def _click_element(self, **kwargs) -> dict:
        x = kwargs.get("x", 0)
        y = kwargs.get("y", 0)
        try:
            if sys.platform == "win32":
                import ctypes
                ctypes.windll.user32.SetCursorPos(x, y)
                ctypes.windll.user32.mouse_event(0x0002, 0, 0, 0, 0)
                ctypes.windll.user32.mouse_event(0x0004, 0, 0, 0, 0)
            return {"success": True, "x": x, "y": y}
        except Exception as e:
            return {"success": False, "error": str(e)}

    async def _type_text(self, **kwargs) -> dict:
        text = kwargs.get("text", "")
        try:
            if sys.platform == "win32":
                import ctypes
                for ch in text:
                    if ch == "\n":
                        ctypes.windll.user32.keybd_event(0x0D, 0, 0, 0)
                        ctypes.windll.user32.keybd_event(0x0D, 0, 0x0002, 0)
                    else:
                        vk = ctypes.windll.user32.VkKeyScanW(ord(ch))
                        ctypes.windll.user32.keybd_event(vk & 0xFF, 0, 0, 0)
                        ctypes.windll.user32.keybd_event(vk & 0xFF, 0, 0x0002, 0)
            return {"success": True, "length": len(text)}
        except Exception as e:
            return {"success": False, "error": str(e)}

    async def _detect_elements(self, **kwargs) -> dict:
        return {"success": True, "elements": [], "message": "Element detection requires vision model"}

    async def _read_screen(self, **kwargs) -> dict:
        try:
            scr = await self._take_screenshot()
            if scr["success"]:
                result = {"success": True, "screenshot": scr["output_path"]}
                return result
            return scr
        except Exception as e:
            return {"success": False, "error": str(e)}


class CodeOfficeSkill(BaseSkill):
    category = SkillCategory.CODE

    def __init__(self):
        self.manifest = SkillManifest(
            name="code-office",
            version="0.1.0",
            category=SkillCategory.CODE,
            level=SkillLevel.INTERMEDIATE,
            description="Code generation, document processing, file operations office suite",
            tags=["code", "office", "document", "file", "generation"],
        )
        self._workspace: Optional[Path] = None

    async def setup(self):
        self._workspace = Path(os.environ.get("COGU_WORKSPACE", os.getcwd()))

    async def execute(self, **kwargs) -> dict:
        action = kwargs.get("action", "generate_code")
        result = {"action": action, "success": False}

        handlers = {
            "generate_code": self._generate_code,
            "analyze_code": self._analyze_code,
            "create_document": self._create_document,
            "convert_format": self._convert_format,
            "list_files": self._list_files,
            "search_files": self._search_files,
            "run_script": self._run_script,
            "install_package": self._install_package,
        }
        handler = handlers.get(action)
        if handler:
            try:
                result.update(await handler(**kwargs))
            except Exception as e:
                result["error"] = str(e)
                result["traceback"] = traceback.format_exc()
        else:
            result["error"] = f"Unknown action: {action}"

        return result

    async def _generate_code(self, **kwargs) -> dict:
        language = kwargs.get("language", "python")
        spec = kwargs.get("spec", "")
        output_file = kwargs.get("output_file", "")
        template = kwargs.get("template", "")

        templates = {
            "python_script": 'import sys\n\ndef main():\n    print("Hello from COGU")\n\nif __name__ == "__main__":\n    main()',
            "python_fastapi": 'from fastapi import FastAPI\n\napp = FastAPI()\n\n@app.get("/")\ndef root():\n    return {"status": "ok"}',
            "html_page": '<!DOCTYPE html>\n<html lang="en">\n<head>\n    <meta charset="UTF-8">\n    <title>COGU Page</title>\n</head>\n<body>\n    <h1>COGU Agent</h1>\n</body>\n</html>',
            "react_component": 'import React from "react";\n\nexport default function App() {\n    return <div>COGU Agent</div>;\n}',
        }

        code = templates.get(template, f"# Generated by COGU\n# {spec}\n")
        if output_file:
            path = self._workspace / output_file if self._workspace else Path(output_file)
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(code, encoding="utf-8")
            return {"success": True, "output_file": str(path), "language": language}
        return {"success": True, "code": code, "language": language}

    async def _analyze_code(self, **kwargs) -> dict:
        file_path = kwargs.get("file_path", "")
        if not file_path:
            return {"success": False, "error": "file_path required"}
        path = Path(file_path)
        if not path.exists():
            return {"success": False, "error": f"File not found: {file_path}"}
        content = path.read_text(encoding="utf-8", errors="replace")
        lines = content.split("\n")
        return {
            "success": True,
            "file": str(path),
            "lines": len(lines),
            "size_bytes": len(content),
            "extension": path.suffix,
            "preview": "\n".join(lines[:20]),
        }

    async def _create_document(self, **kwargs) -> dict:
        doc_type = kwargs.get("doc_type", "markdown")
        title = kwargs.get("title", "Untitled")
        content = kwargs.get("content", "")
        output_file = kwargs.get("output_file", "")

        ext_map = {"markdown": ".md", "text": ".txt", "html": ".html", "json": ".json", "csv": ".csv"}
        ext = ext_map.get(doc_type, ".txt")

        if not output_file:
            output_file = f"{title.replace(' ', '_').lower()}{ext}"

        path = self._workspace / output_file if self._workspace else Path(output_file)
        path.parent.mkdir(parents=True, exist_ok=True)

        if doc_type == "markdown":
            doc_content = f"# {title}\n\n{content}"
        elif doc_type == "html":
            doc_content = f"<!DOCTYPE html>\n<html>\n<head><title>{title}</title></head>\n<body>\n{content}\n</body>\n</html>"
        else:
            doc_content = content

        path.write_text(doc_content, encoding="utf-8")
        return {"success": True, "output_file": str(path), "doc_type": doc_type}

    async def _convert_format(self, **kwargs) -> dict:
        input_file = kwargs.get("input_file", "")
        output_format = kwargs.get("output_format", "txt")

        if not input_file:
            return {"success": False, "error": "input_file required"}

        path = Path(input_file)
        if not path.exists():
            return {"success": False, "error": "File not found: " + input_file}

        content = path.read_text(encoding="utf-8", errors="replace")
        out_path = path.with_suffix(f".{output_format}")
        out_path.write_text(content, encoding="utf-8")

        return {"success": True, "input": str(path), "output": str(out_path), "format": output_format}

    async def _list_files(self, **kwargs) -> dict:
        directory = kwargs.get("directory", str(self._workspace or "."))
        pattern = kwargs.get("pattern", "*")
        recursive = kwargs.get("recursive", False)

        p = Path(directory)
        if not p.exists():
            return {"success": False, "error": f"Directory not found: {directory}"}

        if recursive:
            files = [str(f.relative_to(p)) for f in p.rglob(pattern) if f.is_file()]
        else:
            files = [str(f.name) for f in p.glob(pattern) if f.is_file()]

        return {"success": True, "directory": str(p), "count": len(files), "files": files[:200]}

    async def _search_files(self, **kwargs) -> dict:
        directory = kwargs.get("directory", str(self._workspace or "."))
        query = kwargs.get("query", "")
        file_pattern = kwargs.get("file_pattern", "*")

        if not query:
            return {"success": False, "error": "query required"}

        p = Path(directory)
        results = []
        for f in p.rglob(file_pattern):
            if f.is_file() and f.suffix in (".py", ".md", ".txt", ".json", ".yaml", ".yml", ".toml", ".html", ".js", ".ts", ".css"):
                try:
                    content = f.read_text(encoding="utf-8", errors="replace")
                    if query.lower() in content.lower():
                        results.append(str(f.relative_to(p)))
                except Exception:
                    pass

        return {"success": True, "query": query, "count": len(results), "files": results[:100]}

    async def _run_script(self, **kwargs) -> dict:
        script = kwargs.get("script", "")
        language = kwargs.get("language", "python")
        timeout = kwargs.get("timeout", 30)

        try:
            if language == "python":
                proc = await asyncio.create_subprocess_exec(
                    sys.executable, "-c", script,
                    stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.PIPE,
                    cwd=str(self._workspace) if self._workspace else None,
                )
                stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)
                return {
                    "success": proc.returncode == 0,
                    "stdout": stdout.decode("utf-8", errors="replace"),
                    "stderr": stderr.decode("utf-8", errors="replace"),
                    "returncode": proc.returncode,
                }
            else:
                return {"success": False, "error": f"Language not supported: {language}"}
        except asyncio.TimeoutError:
            return {"success": False, "error": f"Script timed out after {timeout}s"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    async def _install_package(self, **kwargs) -> dict:
        package = kwargs.get("package", "")
        if not package:
            return {"success": False, "error": "package name required"}
        try:
            proc = await asyncio.create_subprocess_exec(
                sys.executable, "-m", "pip", "install", package,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=120)
            return {
                "success": proc.returncode == 0,
                "package": package,
                "stdout": stdout.decode("utf-8", errors="replace"),
            }
        except Exception as e:
            return {"success": False, "error": str(e)}


class BrowserAutomationSkill(BaseSkill):
    category = SkillCategory.BROWSER

    def __init__(self):
        self.manifest = SkillManifest(
            name="browser-automation",
            version="0.1.0",
            category=SkillCategory.BROWSER,
            level=SkillLevel.INTERMEDIATE,
            description="Browser automation via Playwright: navigate, click, fill forms, extract data, screenshots",
            tags=["browser", "playwright", "automation", "scraping", "web"],
            requires_browser=True,
        )

    async def execute(self, **kwargs) -> dict:
        action = kwargs.get("action", "navigate")
        url = kwargs.get("url", "")
        result = {"action": action, "success": False}

        try:
            from playwright.async_api import async_playwright
        except ImportError:
            return {"success": False, "error": "playwright not installed. Run: pip install playwright && playwright install"}

        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=True)
            page = await browser.new_page()

            try:
                if action == "navigate":
                    if url:
                        await page.goto(url, wait_until="networkidle")
                        result.update({"success": True, "title": await page.title(), "url": page.url})
                elif action == "screenshot":
                    if url:
                        await page.goto(url, wait_until="networkidle")
                    output = kwargs.get("output_path", "screenshot.png")
                    await page.screenshot(path=output, full_page=True)
                    result.update({"success": True, "output_path": output})
                elif action == "extract_text":
                    if url:
                        await page.goto(url, wait_until="networkidle")
                    selector = kwargs.get("selector", "body")
                    text = await page.text_content(selector)
                    result.update({"success": True, "text": text[:5000]})
                elif action == "fill_form":
                    if url:
                        await page.goto(url, wait_until="networkidle")
                    fields = kwargs.get("fields", {})
                    for sel, val in fields.items():
                        await page.fill(sel, str(val))
                    result.update({"success": True, "filled": list(fields.keys())})
                elif action == "click":
                    if url:
                        await page.goto(url, wait_until="networkidle")
                    selector = kwargs.get("selector", "")
                    if selector:
                        await page.click(selector)
                        result.update({"success": True, "clicked": selector})
                elif action == "evaluate":
                    script = kwargs.get("script", "")
                    if script:
                        output = await page.evaluate(script)
                        result.update({"success": True, "result": str(output)[:2000]})
            except Exception as e:
                result["error"] = str(e)
            finally:
                await browser.close()

        return result


class DataAnalysisSkill(BaseSkill):
    category = SkillCategory.DATA

    def __init__(self):
        self.manifest = SkillManifest(
            name="data-analysis",
            version="0.1.0",
            category=SkillCategory.DATA,
            level=SkillLevel.INTERMEDIATE,
            description="Data analysis: CSV/JSON/Excel parsing, statistics, visualization, transformation",
            tags=["data", "analysis", "csv", "excel", "statistics", "visualization"],
        )

    async def execute(self, **kwargs) -> dict:
        action = kwargs.get("action", "analyze")
        result = {"action": action, "success": False}

        try:
            if action == "read_csv":
                result.update(self._read_csv(**kwargs))
            elif action == "read_excel":
                result.update(self._read_excel(**kwargs))
            elif action == "read_json":
                result.update(self._read_json(**kwargs))
            elif action == "stats":
                result.update(self._compute_stats(**kwargs))
            elif action == "filter":
                result.update(self._filter_data(**kwargs))
            elif action == "to_csv":
                result.update(self._to_csv(**kwargs))
            elif action == "to_excel":
                result.update(self._to_excel(**kwargs))
            else:
                result["error"] = f"Unknown action: {action}"
        except Exception as e:
            result["error"] = str(e)

        return result

    def _read_csv(self, **kwargs) -> dict:
        import csv
        path = kwargs.get("file_path", "")
        if not path:
            return {"success": False, "error": "file_path required"}
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        return {"success": True, "rows": len(rows), "columns": reader.fieldnames, "preview": rows[:20]}

    def _read_excel(self, **kwargs) -> dict:
        path = kwargs.get("file_path", "")
        sheet = kwargs.get("sheet", 0)
        if not path:
            return {"success": False, "error": "file_path required"}
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            wb.close()
            return {"success": True, "rows": len(rows), "columns": headers, "preview": rows[:20]}
        except ImportError:
            return {"success": False, "error": "openpyxl not installed"}

    def _read_json(self, **kwargs) -> dict:
        path = kwargs.get("file_path", "")
        if not path:
            return {"success": False, "error": "file_path required"}
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return {"success": True, "type": "array", "count": len(data), "preview": data[:10]}
        return {"success": True, "type": "object", "keys": list(data.keys())[:50]}

    def _compute_stats(self, **kwargs) -> dict:
        data = kwargs.get("data", [])
        column = kwargs.get("column", "")
        if not data or not column:
            return {"success": False, "error": "data and column required"}
        values = [float(row.get(column, 0)) for row in data if row.get(column) is not None]
        if not values:
            return {"success": False, "error": "No numeric values found"}
        values.sort()
        n = len(values)
        return {
            "success": True,
            "count": n,
            "min": min(values),
            "max": max(values),
            "mean": sum(values) / n,
            "median": values[n // 2],
            "sum": sum(values),
        }

    def _filter_data(self, **kwargs) -> dict:
        data = kwargs.get("data", [])
        filters = kwargs.get("filters", {})
        if not data:
            return {"success": False, "error": "data required"}
        filtered = []
        for row in data:
            match = True
            for col, val in filters.items():
                if str(row.get(col, "")) != str(val):
                    match = False
                    break
            if match:
                filtered.append(row)
        return {"success": True, "original": len(data), "filtered": len(filtered), "data": filtered[:100]}

    def _to_csv(self, **kwargs) -> dict:
        import csv
        data = kwargs.get("data", [])
        output = kwargs.get("output_path", "output.csv")
        if not data:
            return {"success": False, "error": "data required"}
        keys = list(data[0].keys())
        with open(output, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            writer.writerows(data)
        return {"success": True, "output_path": output, "rows": len(data)}

    def _to_excel(self, **kwargs) -> dict:
        data = kwargs.get("data", [])
        output = kwargs.get("output_path", "output.xlsx")
        if not data:
            return {"success": False, "error": "data required"}
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            keys = list(data[0].keys())
            ws.append(keys)
            for row in data:
                ws.append([row.get(k) for k in keys])
            wb.save(output)
            return {"success": True, "output_path": output, "rows": len(data)}
        except ImportError:
            return {"success": False, "error": "openpyxl not installed"}


class WebSearchSkill(BaseSkill):
    category = SkillCategory.DATA

    def __init__(self):
        self.manifest = SkillManifest(
            name="web-search",
            version="0.1.0",
            category=SkillCategory.DATA,
            level=SkillLevel.BASIC,
            description="Web search via DuckDuckGo: query, news, extract page content",
            tags=["web", "search", "internet", "news"],
        )

    async def execute(self, **kwargs) -> dict:
        action = kwargs.get("action", "search")
        result = {"action": action, "success": False}

        try:
            from cogu.tools.builtin.web import _web_search, _web_fetch, _web_news
        except ImportError:
            return {"success": False, "error": "Web tools not available"}

        try:
            if action == "search":
                query = kwargs.get("query", "")
                num = kwargs.get("num", 10)
                out = _web_search(query, num)
                result.update({"success": True, "results": out})
            elif action == "fetch":
                url = kwargs.get("url", "")
                out = _web_fetch(url)
                result.update({"success": True, "content": str(out)[:10000]})
            elif action == "news":
                topic = kwargs.get("topic", "")
                out = _web_news(topic)
                result.update({"success": True, "results": out})
            elif action == "fetch_json":
                url = kwargs.get("url", "")
                import urllib.request, json
                with urllib.request.urlopen(url, timeout=15) as resp:
                    data = json.loads(resp.read())
                result.update({"success": True, "data": data})
            else:
                result["error"] = f"Unknown action: {action}"
        except Exception as e:
            result["error"] = str(e)

        return result


class ShellSkill(BaseSkill):
    category = SkillCategory.CODE

    def __init__(self):
        self.manifest = SkillManifest(
            name="shell",
            version="0.1.0",
            category=SkillCategory.CODE,
            level=SkillLevel.BASIC,
            description="Shell command execution: run, pipe, file ops, process management",
            tags=["shell", "terminal", "command", "system"],
        )

    async def execute(self, **kwargs) -> dict:
        action = kwargs.get("action", "run")
        result = {"action": action, "success": False}

        try:
            if action == "run":
                command = kwargs.get("command", "")
                timeout = kwargs.get("timeout", 30)
                cwd = kwargs.get("cwd", None)
                proc = await asyncio.create_subprocess_shell(
                    command,
                    stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.PIPE,
                    cwd=cwd,
                )
                stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)
                result.update({
                    "success": proc.returncode == 0,
                    "stdout": stdout.decode("utf-8", errors="replace"),
                    "stderr": stderr.decode("utf-8", errors="replace"),
                    "returncode": proc.returncode,
                })
            elif action == "exec":
                program = kwargs.get("program", "")
                args = kwargs.get("args", [])
                timeout = kwargs.get("timeout", 30)
                proc = await asyncio.create_subprocess_exec(
                    program, *args,
                    stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.PIPE,
                )
                stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)
                result.update({
                    "success": proc.returncode == 0,
                    "stdout": stdout.decode("utf-8", errors="replace"),
                    "stderr": stderr.decode("utf-8", errors="replace"),
                    "returncode": proc.returncode,
                })
            elif action == "which":
                program = kwargs.get("program", "")
                import shutil
                path = shutil.which(program)
                result.update({"success": True, "found": path is not None, "path": path})
            else:
                result["error"] = f"Unknown action: {action}"
        except asyncio.TimeoutError:
            result["error"] = f"Command timed out after {kwargs.get('timeout', 30)}s"
        except Exception as e:
            result["error"] = str(e)

        return result


class FinanceSkill(BaseSkill):
    category = SkillCategory.DATA

    def __init__(self):
        self.manifest = SkillManifest(
            name="finance",
            version="0.1.0",
            category=SkillCategory.DATA,
            level=SkillLevel.INTERMEDIATE,
            description="Financial data queries: stock quotes, forex, crypto prices",
            tags=["finance", "stock", "forex", "crypto", "market"],
        )

    async def execute(self, **kwargs) -> dict:
        action = kwargs.get("action", "quote")
        result = {"action": action, "success": False}

        try:
            from cogu.tools.builtin.stock import _get_stock_quote, _get_forex_rate, _get_crypto_price
        except ImportError:
            return {"success": False, "error": "Stock tools not available"}

        try:
            if action == "stock":
                symbol = kwargs.get("symbol", "")
                out = _get_stock_quote(symbol)
                result.update({"success": True, "quote": out})
            elif action == "forex":
                pair = kwargs.get("pair", "USD/CNY")
                out = _get_forex_rate(pair)
                result.update({"success": True, "rate": out})
            elif action == "crypto":
                coin = kwargs.get("coin", "BTC")
                out = _get_crypto_price(coin)
                result.update({"success": True, "price": out})
            elif action == "search":
                query = kwargs.get("query", "")
                result.update({"success": True, "query": query, "message": "Use web-search skill for financial news search"})
            else:
                result["error"] = f"Unknown action: {action}"
        except Exception as e:
            result["error"] = str(e)

        return result


class GatewaySkill(BaseSkill):
    category = SkillCategory.COMMUNICATION

    def __init__(self):
        self.manifest = SkillManifest(
            name="gateway",
            version="0.1.0",
            category=SkillCategory.COMMUNICATION,
            level=SkillLevel.INTERMEDIATE,
            description="API gateway management: start, stop, status, route info",
            tags=["gateway", "api", "server", "management"],
        )

    async def execute(self, **kwargs) -> dict:
        action = kwargs.get("action", "status")
        result = {"action": action, "success": False}

        try:
            if action == "start":
                port = kwargs.get("port", 8080)
                host = kwargs.get("host", "127.0.0.1")
                try:
                    from cogu.gateway.server import GatewayServer
                    self._gateway = GatewayServer(host=host, port=port)
                    await self._gateway.start()
                    result.update({"success": True, "host": host, "port": port, "status": "running"})
                except ImportError:
                    result.update({"success": True, "message": f"Gateway module loaded. Use 'cogu serve --port {port}' to start"})
            elif action == "stop":
                if hasattr(self, '_gateway'):
                    await self._gateway.stop()
                    result.update({"success": True, "status": "stopped"})
                else:
                    result.update({"success": True, "status": "not running"})
            elif action == "status":
                result.update({"success": True, "status": "checking", "message": "Gateway status available via cogu serve command"})
            elif action == "routes":
                try:
                    from cogu.app import create_app
                    app = create_app()
                    routes = [{"path": r.path, "methods": list(r.methods)} for r in app.routes]
                    result.update({"success": True, "routes": routes})
                except ImportError:
                    result.update({"success": True, "routes": [], "message": "App module not loaded"})
            else:
                result["error"] = f"Unknown action: {action}"
        except Exception as e:
            result["error"] = str(e)

        return result


class PromptSkill(BaseSkill):
    def __init__(self, name: str, description: str = "", version: str = "0.1.0",
                 tags: list = None, prompt_content: str = "", skill_dir: str = ""):
        self.manifest = SkillManifest(
            name=name,
            version=version,
            category=SkillCategory.CUSTOM,
            level=SkillLevel.BASIC,
            description=description,
            tags=tags or [],
        )
        self._prompt_content = prompt_content
        self._skill_dir = skill_dir

    async def execute(self, **kwargs) -> dict:
        return {
            "success": True,
            "skill_type": "prompt",
            "name": self.manifest.name,
            "prompt": self._prompt_content,
            "skill_dir": self._skill_dir,
            "message": "Prompt skill: pass this content to LLM as context",
        }


class BuiltinSkillRegistry:
    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._skills: dict[str, BaseSkill] = {}
            cls._instance._initialized = False
        return cls._instance

    def register(self, skill: BaseSkill):
        self._skills[skill.manifest.name] = skill

    def unregister(self, name: str):
        self._skills.pop(name, None)

    def get(self, name: str) -> Optional[BaseSkill]:
        return self._skills.get(name)

    def list_by_category(self, category: SkillCategory) -> list[BaseSkill]:
        return [s for s in self._skills.values() if s.manifest.category == category]

    def list_all(self) -> list[BaseSkill]:
        return list(self._skills.values())

    def describe_all(self) -> list[dict]:
        return [s.describe() for s in self._skills.values()]

    async def execute(self, skill_name: str, **kwargs) -> dict:
        skill = self._skills.get(skill_name)
        if not skill:
            return {"success": False, "error": f"Skill not found: {skill_name}"}
        return await skill.execute(**kwargs)

    async def initialize(self):
        if self._initialized:
            return

        builtin_skills = [
            ReasoningSkill(),
            GuiVisionSkill(),
            CodeOfficeSkill(),
            BrowserAutomationSkill(),
            DataAnalysisSkill(),
            WebSearchSkill(),
            ShellSkill(),
            FinanceSkill(),
            GatewaySkill(),
        ]
        for skill in builtin_skills:
            self.register(skill)
            await skill.setup()

        loop_skills = self._load_loop_skills()
        for skill in loop_skills:
            self.register(skill)
            await skill.setup()

        bundled_dir = Path(__file__).resolve().parent.parent.parent / "skills"
        if bundled_dir.is_dir():
            await self.load_from_directory(bundled_dir)

        self._initialized = True

    def _load_loop_skills(self) -> list[BaseSkill]:
        try:
            from cogu.skills.loop_triage import LoopTriageSkill
            from cogu.skills.loop_verifier import LoopVerifierSkill
            from cogu.skills.loop_budget import LoopBudgetSkill
            from cogu.skills.minimal_fix import MinimalFixSkill
            return [
                LoopTriageSkill(),
                LoopVerifierSkill(),
                LoopBudgetSkill(),
                MinimalFixSkill(),
            ]
        except ImportError:
            return []

    async def load_from_directory(self, directory: Path | str):
        dir_path = Path(directory)
        if not dir_path.exists():
            return
        for py_file in dir_path.glob("*.py"):
            try:
                spec = importlib.util.spec_from_file_location(py_file.stem, py_file)
                module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(module)
                for name, obj in inspect.getmembers(module, inspect.isclass):
                    if issubclass(obj, BaseSkill) and obj is not BaseSkill:
                        skill = obj()
                        self.register(skill)
                        await skill.setup()
            except Exception:
                pass
        for skill_md in dir_path.rglob("SKILL.md"):
            skill_dir = skill_md.parent
            if skill_dir.name.startswith(('.', '_')):
                continue
            try:
                self._load_skill_md(skill_dir, skill_md)
            except Exception:
                pass
        for meta_json in dir_path.rglob("_meta.json"):
            skill_dir = meta_json.parent
            if skill_dir.name.startswith(('.', '_')):
                continue
            skill_md = skill_dir / "SKILL.md"
            if skill_md.exists():
                continue
            try:
                self._load_meta_skill(skill_dir, meta_json)
            except Exception:
                pass

    def _load_skill_md(self, skill_dir: Path, skill_md: Path):
        content = skill_md.read_text(encoding="utf-8")
        name = skill_dir.name
        description = ""
        tags = []
        version = "0.1.0"
        if content.startswith("---"):
            parts = content.split("---", 2)
            if len(parts) >= 3:
                front = parts[1].strip()
                for line in front.split("\n"):
                    if line.startswith("name:"):
                        name = line.split(":", 1)[1].strip().strip('"')
                    elif line.startswith("description:"):
                        description = line.split(":", 1)[1].strip().strip('"')
                    elif line.startswith("tags:"):
                        tags = [t.strip() for t in line.split(":", 1)[1].split(",")]
                    elif line.startswith("version:"):
                        version = line.split(":", 1)[1].strip()

        skill = PromptSkill(
            name=name,
            description=description,
            version=version,
            tags=tags,
            prompt_content=content,
            skill_dir=str(skill_dir),
        )
        self.register(skill)

    def _load_meta_skill(self, skill_dir: Path, meta_json: Path):
        data = json.loads(meta_json.read_text(encoding="utf-8"))
        name = data.get("name", skill_dir.name)
        description = data.get("description", "")
        version = data.get("version", "0.1.0")
        tags = data.get("tags", [])
        prompt_content = ""
        for fname in ("SKILL.md", "README.md"):
            fpath = skill_dir / fname
            if fpath.exists():
                prompt_content = fpath.read_text(encoding="utf-8")
                break
        skill = PromptSkill(
            name=name,
            description=description,
            version=version,
            tags=tags,
            prompt_content=prompt_content,
            skill_dir=str(skill_dir),
        )
        self.register(skill)

    @classmethod
    def instance(cls):
        return cls()


def get_builtin_skill_registry() -> BuiltinSkillRegistry:
    return BuiltinSkillRegistry.instance()
