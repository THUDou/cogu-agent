import json
import re
import sys
from typing import Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string

CELL_RE = re.compile(r"^([A-Z]+)([0-9]+)$")


def parse_cell(cell_ref: str) -> Tuple[int, int]:
    m = CELL_RE.match(cell_ref.strip().upper())
    if not m:
        raise ValueError(f"Invalid cell reference: {cell_ref}")
    col_str, row_str = m.groups()
    return int(row_str), column_index_from_string(col_str)


def parse_range(start: str, end: Optional[str] = None) -> Tuple[int, int, int, int]:
    if start and ":" in start and end is None:
        start, end = start.split(":", 1)
    sr, sc = parse_cell(start)
    if end:
        er, ec = parse_cell(end)
    else:
        er, ec = sr, sc
    return sr, sc, er, ec


def load_or_create_wb(filepath: str) -> Workbook:
    try:
        return load_workbook(filepath)
    except FileNotFoundError:
        return Workbook()


def require_sheet(wb: Workbook, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    return wb[sheet_name]


def emit(payload: dict, exit_code: int = 0):
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    sys.exit(exit_code)


def emit_error(message: str, exit_code: int = 1, **extra):
    emit({"status": "error", "error": message, **extra}, exit_code=exit_code)
